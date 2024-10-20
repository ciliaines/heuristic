# This set of functions is for the visualization of the values of the ILP 
import pandas as pd
from openpyxl import load_workbook 

from RanNet_Generator import *
from Djikstra_Path_Calculator import *
from RandStream_Parameters import *
from Preprocessing import *
from Heuristic_Generator import *
import time
from read import *
from write import *
import textwrap
from Plot import *
from datetime import datetime


now = datetime.now()
timestamp = now.strftime("%Y-%m-%d_%H:%M:%S")

input = "input1"
input_timestamp = "input1_" + timestamp
input_name = input + "_heuristic_" + timestamp
file_input_topo = "Inputs/" + input + "_topo.json"
file_input = "Solutions/" + input_timestamp + ".json"
file_result = "Results/"+ input + "_result_" + ".json"
file_image = "Solutions/" + input_name + ".html"

Hyperperiod = 1000
#Hyperperiod = 6000
#Hyperperiod = 30000

def Heuristic_results_visualizer(instance, Model_Descriptor_vector):
    print("############### This is the set of offsets ######################")
    Result_offsets = []
    Clean_offsets_collector = []
    Feasibility_indicator = 0
    for i in instance.Streams:
        for j in instance.Links:
            for k in instance.Frames:
                if Model_Descriptor_vector [i][k][j] :
                    print("The offset of stream", i, "link", j, "frame", k, "is",instance.Lower_bound[i,j,k].value)
                    frame_indicator = ("S", i, "L", j, "F", k, "Q",int(instance.Queue_Assignment[i, j].value), "La",int(instance.Latency[i].value), "O",int(instance.Frame_Offset[i,j,k].value))
                    helper = { "Task" :str(frame_indicator), "Start": instance.Lower_bound[i,j,k].value, "Finish" : instance.Upper_bound[i,j,k].value, "Color" : j }
                    clean_offset = { "Task" :str(frame_indicator), "Start": instance.Lower_bound[i,j,k].value }
                    Result_offsets.append(helper)
                    Clean_offsets_collector.append(clean_offset)
                    if instance.Frame_Offset[i,j,k].value != 1 :
                        Feasibility_indicator = Feasibility_indicator + 1         
    print("############### This is the set of latencies ######################")
    Results_latencies = []
    for stream in instance.Streams:
        #print("The lower latency of Stream", stream, "is",instance.Lower_Latency[stream].value)
        print("The latency of Stream ", stream, "is", instance.Latency[stream].value)
        Results_latencies.append(instance.Latency[stream].value)#
    for i in instance.Streams:
        for j in instance.Links:
            instance.Num_Queues[j] = max(instance.Num_Queues[j].value , instance.Queue_Assignment[i,j].value)
    print("############### This is the set of queues ######################")
    for link in instance.Links:
        print("The number of queues of link ", link, "is", instance.Num_Queues[link].value+1)

    print("############### This is the set of queues per stream and link######################")
    for stream in instance.Streams:
        for link in instance.Links:
            print("The number of queues of Link",link , "Stream" , stream, "is", instance.Queue_Assignment[stream, link].value)

#    print("############### This is the set of auxiliar queues variables######################")
#    for stream in instance.Streams :
#        for stream_2 in instance.Streams :
#            for link in instance.Links :
#                print("Aux variable for stream_1 ", stream, "Stream_2", stream_2, "link", link, ":", instance.Aux_Same_Queue[stream_2, link ,stream].value)
    return Feasibility_indicator, Result_offsets, Clean_offsets_collector, Results_latencies

##### For printing the model results and variables #####
#UNCOMMENT if necessary
# instance.display()
# results.write()
# results.solver.status 
######################## For now on, this code is for generate the Gant chart ########################

def Evaluation_function(Number_of_edges, Connection_probability,Number_of_Streams) :
### This is just the part where the program can select betweeen generating a new network
################################################################
    Stream_Source_Destination = []
    Streams_Period = {} 
    Deathline_Stream = {}
    Streams_size = list()
    Number_of_Streams=0
    try :
        initial_time = time.time()
        utilizacion = True
        num_stream = 0
        while utilizacion:
            #Read from the JSON files
            if num_stream == 0:
                Number_of_edges, Number_of_Streams, Network_nodes, Network_links, Adjacency_Matrix, plot_network, Sources, Destinations, Stream_Source_Destination_total = Read(file_input_topo)
            Stream_Source_Destination,Streams_Period, Deathline_Stream, Number_of_Streams, Streams_size = Random(Stream_Source_Destination_total, Hyperperiod, Stream_Source_Destination, Streams_Period, Deathline_Stream, Number_of_Streams, Streams_size)
            ################################################################
            #Djikstra scheduler
            network = Network_Topology(Adjacency_Matrix) 
            all_paths_matrix = all_paths_matrix_generator(Network_nodes, network)
            Streams_paths = Streams_paths_generator(all_paths_matrix, Stream_Source_Destination)
            Streams_links_paths = Streams_links_paths_generator(Streams_paths)
            Link_order_Descriptor = Link_order_Descriptor_generator(Streams_links_paths, Network_links)
            ###############################################################
            # Random Streams parameters
            Sort_Stream_Source_Destination = Sort_flow(Stream_Source_Destination, Deathline_Stream, Streams_Period, Streams_size)
            Frames_per_Stream, Max_frames, Num_of_Frames = Frames_per_Stream_generator(Streams_size)
            ################################################################
            # Preprocessing
            Links_per_Stream = Links_per_Stream_generator(Network_links, Link_order_Descriptor)
            Model_Descriptor, Model_Descriptor_vector, Streams = Model_Descriptor_generator(Number_of_Streams, Max_frames, Network_links, Frames_per_Stream, Links_per_Stream)
            Frame_Duration =  Frame_Duration_Generator(Number_of_Streams, Max_frames, Network_links )
            Repetitions, Repetitions_Matrix, Repetitions_Descriptor, max_repetitions= Repetitions_generator(Streams_Period, Streams, Hyperperiod)
            ################################################################
            scheduler = Heuristic_class(Number_of_Streams, Network_links, \
                    Link_order_Descriptor, \
                    Streams_Period, Hyperperiod, Frames_per_Stream, Max_frames, Num_of_Frames, \
                    Model_Descriptor, Model_Descriptor_vector, Deathline_Stream, \
                    Repetitions, Repetitions_Descriptor, Frame_Duration, \
                    Stream_Source_Destination, Streams_size, Streams_paths, Sort_Stream_Source_Destination) #unused_links¿?

            instance, results = scheduler.instance, scheduler.results
            utilizacion = Greedy_Heuristic(instance, num_stream)
            num_stream = num_stream + 1

        
        final_time = time.time()        ### This will store the results into a txt for further usage
        time_evaluation = final_time - initial_time
        
        ################################################################
        Feasibility_indicator, Result_offsets, Clean_offsets_collector, Results_latencies = Heuristic_results_visualizer(instance, Model_Descriptor_vector)
        Write(input, input_timestamp,Number_of_Streams,Streams_Period,Deathline_Stream,Streams_size,Stream_Source_Destination)
        set_offset=""
        lower_latency=""
        queues_link=""
        queues_stream=""

        file_path_time = 'heuristico_time.xlsx'
        file_path_offset = 'heuristico_offset.xlsx'
        file_path_latencies = 'heuristico_latencies.xlsx'
        file_path_queues = 'heuristico_queues.xlsx'

         
       # df_existing = pd.read_excel(file_path)
        wb_time = load_workbook(file_path_time)
        wb_offset = load_workbook(file_path_offset)
        wb_latencies = load_workbook(file_path_latencies)
        wb_queues = load_workbook(file_path_queues)

        ws_time = wb_time.active
        ws_offset = wb_offset.active
        ws_latencies =wb_latencies.active
        ws_queues = wb_queues.active
        #guardar varibale del tiempo en un archovp

        with open ('variable.txt', 'w') as file:
            file.write(timestamp)
        
        with open('Results/' + input_name + '.txt', 'a') as f :
            #Time
            f.write("Execution time:    ")
            f.write(str(time_evaluation) + "\n")
            header = ["Experimento", "Time"]
            body = [input_name, time_evaluation]
            row =ws_time.max_row + 1
            ws_time.append(body)
            wb_time.save(file_path_time)

            
            f.write("############### This is the set of offsets ######################" + "\n")       
            for i in instance.Streams:
                for j in instance.Links:
                    for k in instance.Frames:
                        if Model_Descriptor_vector [i][k][j] :
                            f.write("The offset of stream " + str(i) + " link " +str(j)+ " frame " + str(k) + " is " + str(instance.Lower_bound[i,j,k].value) + "\n")
                            set_offset+= str(i) + " for " +str(j)+ " frame "+str(k) + " is " + str(instance.Lower_bound[i,j,k].value) + "<br>"
                            header = ["Experimento", "Stream", "Link", "Frame", "Offset" ]
                            body = [input_name, str(i), str(j),  str(k), str(instance.Lower_bound[i,j,k].value)]
                            row =ws_offset.max_row + 1
                            ws_offset.append(body)
                            wb_offset.save(file_path_offset)

            f.write("############### This is the set of latencies ######################" + "\n")
            for stream in instance.Streams:
                f.write("The lower latency of Stream " + str(stream) + " is " + str(instance.Latency[stream].value) + "\n")
                lower_latency += str(stream) + " is " + str(instance.Latency[stream].value) + "<br>"
                header = ["Experimento", "Stream", "Latency" ]
                body = [input_name, str(stream),str(instance.Latency[stream].value) ]
                row =ws_latencies.max_row + 1
                ws_latencies.append(body)
                wb_latencies.save(file_path_latencies)

            f.write("############### This is the set of queues ######################" + "\n")
            for link in instance.Links:
                f.write("The number of queues of link " + str(link) + " is " + str(instance.Num_Queues[link].value+1) + "\n")
                queues_link += str(link) + " is " + str(instance.Num_Queues[link].value+1) + "<br>"

            f.write("############### This is the set of queues per stream and link######################" + "\n")
            for stream in instance.Streams:
                for link in instance.Links:
                    f.write("The number of queues of Link " + str(link) + " stream " + str(stream) + " is " + str(instance.Queue_Assignment[stream, link].value) + "\n")
                    queues_stream += str(link) + " for " + str(stream) + " is " + str(instance.Queue_Assignment[stream, link].value) + "<br>"
                    header = ["Experimento", "Link", "Stream", "Queue" ]
                    body = [input_name, str(link), str(stream), str(instance.Queue_Assignment[stream, link].value) ]
                    row =ws_queues.max_row + 1
                    ws_queues.append(body)
                    wb_queues.save(file_path_queues)

        

        #PLOT
        network_info_fig = network_info_topology(input_name,Sources,Destinations,Network_links, Repetitions, Streams_Period, Link_order_Descriptor, Streams_links_paths)
        gantt_fig = gantt_chart(Result_offsets, Repetitions, Streams_Period)
        #info_fig = info_box(Network_links, Repetitions, Streams_Period, Link_order_Descriptor, Streams_links_paths)
        result_fig = result_box(time_evaluation, set_offset,lower_latency,queues_link,queues_stream)
        combined(network_info_fig, gantt_fig, result_fig, file_image)
    except ValueError:
        print("One error has occurred")

for n in [4]:
    for i in range(1):
        # Evaluation_Function(number_of_nodes, connection_probability, number of streams)
        Evaluation_function(2, 1, n)
